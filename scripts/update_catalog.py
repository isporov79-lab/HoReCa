from pathlib import Path
from urllib.parse import quote
import hashlib
import re

import requests
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

SOURCE_XLSX = ROOT / "catalog.xlsx"
OUTPUT_XLSX = ROOT / "assets" / "files" / "products.xlsx"
IMAGE_DIR = ROOT / "assets" / "img" / "products"

API_URL = (
    "https://cloud-api.yandex.net/v1/disk/public/resources/download"
    "?public_key="
)

IMAGE_EXTS = {
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
}


def is_yandex_public_link(value):
    if not isinstance(value, str):
        return False

    return bool(
        re.search(
            r"https?://(?:www\.)?disk\.yandex\.(?:ru|com|kz|uz|tr)/i/",
            value.strip(),
            re.IGNORECASE,
        )
    )


def safe_name(text):
    text = re.sub(
        r"[^0-9A-Za-zА-Яа-яЁё_-]+",
        "_",
        str(text),
    ).strip("_")

    return text[:80] or "product"


def yandex_download(public_url):
    api_response = requests.get(
        API_URL + quote(public_url.strip(), safe=""),
        timeout=30,
    )
    api_response.raise_for_status()

    data = api_response.json()
    href = data.get("href")

    if not href:
        raise RuntimeError(
            "Yandex API did not return download href"
        )

    image_response = requests.get(
        href,
        timeout=60,
        stream=True,
    )
    image_response.raise_for_status()

    content_type = (
        image_response.headers.get("Content-Type") or ""
    ).split(";")[0].lower()

    ext = IMAGE_EXTS.get(content_type)

    if not ext:
        ext = Path(
            href.split("?", 1)[0]
        ).suffix.lower()

    if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
        ext = ".jpg"

    return image_response, ext


def main():
    if not SOURCE_XLSX.exists():
        raise SystemExit(
            f"Не найден файл: {SOURCE_XLSX}"
        )

    IMAGE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    OUTPUT_XLSX.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb = load_workbook(SOURCE_XLSX)
    ws = wb[wb.sheetnames[0]]

    headers = {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }

    image_col = headers.get("image")

    if not image_col:
        raise SystemExit(
            "В Excel не найдена колонка image"
        )

    name_col = headers.get("name")
    brand_col = headers.get("brand")

    downloaded = 0
    failed = 0
    skipped = 0

    for row in range(2, ws.max_row + 1):
        value = ws.cell(
            row=row,
            column=image_col,
        ).value

        # Пропускаем пустые значения
        if not value:
            continue

        # Если фото уже локальное — ничего не скачиваем
        if isinstance(value, str) and value.startswith(
            "assets/img/products/"
        ):
            skipped += 1
            continue

        # Обрабатываем только ссылки Yandex Disk
        if not is_yandex_public_link(value):
            continue

        brand = (
            ws.cell(row=row, column=brand_col).value
            if brand_col
            else ""
        )

        name = (
            ws.cell(row=row, column=name_col).value
            if name_col
            else ""
        )

        key = (
            f"{row}-{brand}-{name}-{value}"
            .encode("utf-8", "ignore")
        )

        stem = (
            safe_name(f"{brand}_{name}")
            + "-"
            + hashlib.sha1(key).hexdigest()[:8]
        )

        try:
            response, ext = yandex_download(value)

            try:
                filename = stem + ext
                target = IMAGE_DIR / filename

                with target.open("wb") as f:
                    for chunk in response.iter_content(
                        1024 * 64
                    ):
                        if chunk:
                            f.write(chunk)

            finally:
                response.close()

            # ВАЖНО:
            # именно catalog.xlsx потом читает convert_catalog.py
            local_path = (
                f"assets/img/products/{filename}"
            )

            ws.cell(
                row=row,
                column=image_col,
            ).value = local_path

            downloaded += 1

            print(
                f"OK: {value} -> {local_path}"
            )

        except Exception as exc:
            failed += 1

            print(
                f"ERROR: {value}: {exc}"
            )

    # Сохраняем ОБНОВЛЁННЫЙ catalog.xlsx
    # чтобы convert_catalog.py увидел локальные пути
    wb.save(SOURCE_XLSX)

    # Дополнительно создаём копию
    wb.save(OUTPUT_XLSX)

    print(
        f"Готово. Загружено фото: {downloaded}, "
        f"пропущено: {skipped}, ошибок: {failed}"
    )

    print(
        f"Обновлён: {SOURCE_XLSX}"
    )

    print(
        f"Копия создана: {OUTPUT_XLSX}"
    )


if __name__ == "__main__":
    main()
